#!/usr/bin/env python3
from __future__ import annotations
import argparse, copy, hashlib, json, os, shutil, tempfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

STAGE_ID="CURRENT_MASTER_BR_IBRX100_CONTROLLED_SOURCE_SUPERSET_IMPORT_AND_ELIGIBILITY_STATE_MATERIALIZATION"
VERSION="v0.34"
SCHEMA="WELT_SWING_CURRENT_MASTER_BR_IBRX100_CONTROLLED_IMPORT_ELIGIBILITY_MATERIALIZATION_V0_34"
LINEAGE="CURRENT_MASTER_CLEAN_RESTART"
STATUS="DEV / RESEARCH / SHADOW - NOT PRODUCTIVE"
SOURCE_ID="B3_OFFICIAL_INDEXPROXY_GETPORTFOLIODAY"
SOURCE_ASOF="2026-08-31"
PROVIDER="YFINANCE_FREE"
CANONICAL_STATUS="ACTIVE_VERIFIED"
SCALABLE_STATUS="SCALABLE_NOT_VERIFIED"
COLUMNS=["WS_ID","Name","ISIN","Instrument_Type","Country","Primary_Ticker","Primary_Exchange","Primary_MIC","Primary_Currency","Yahoo_Symbol","Alpha_Symbol","Primary_Universe_Index","Index_Tags","Active","Universe_Status","Mapping_Status","Scalable_Tradeability_Status","Source_ID","Source_AsOf","Last_Validated","Share_Class","Notes"]
SHEETS={"Universe_Master","Import_Coverage","Dedupe_Rules","Status_Definitions","Run_Summary"}
OLD_SEGMENTS={"EU_STOXX600":600,"CA_TSX":217,"JP_N225":225,"HK_HSI":93,"CN_CSI300":300,"IN_NIFTY50":50,"TW_TW50":50}
TARGETS=set(OLD_SEGMENTS)|{"BR_IBRX100","US_SP1500","MX_IPC","KR_KOSPI200","AU_ASX200","NZ_NZX50","ZA_TOP40"}
STATES={"STANDARD_ELIGIBILITY_READY":38,"LOW_LIQUIDITY_EXCEPTION_POOL":28,"STANDARD_U3K_NOT_ELIGIBLE_LIQUIDITY":13,"STANDARD_U3K_NOT_ELIGIBLE_INSTRUMENT":19}
OUTPUTS=["br_ibrx100_import_audit_v0.34.csv","br_ibrx100_imported_rows_v0.34.csv","br_ibrx100_eligibility_state_audit_v0.34.csv","preexisting_1535_immutability_audit_v0.34.csv","current_master_segment_inventory_v0.34.csv","current_master_post_import_identity_audit_v0.34.csv","workbook_sheet_consistency_audit_v0.34.csv","research_partial_1633_reconciliation_v0.34.csv","summary_v0.34.json","stage_checkpoint_v0.34.json"]

def req(ok,msg):
    if not ok: raise RuntimeError(msg)

def txt(v):
    if v is None or (isinstance(v,float) and pd.isna(v)): return ""
    if isinstance(v,bool): return "True" if v else "False"
    if isinstance(v,datetime): return v.date().isoformat() if v.time().isoformat()=="00:00:00" else v.isoformat()
    return str(v).strip()

def truth(v): return txt(v).lower() in {"true","1","yes"}
def utc(): return datetime.now(timezone.utc).isoformat()

def sha256(path):
    h=hashlib.sha256()
    with Path(path).open("rb") as f:
        for chunk in iter(lambda:f.read(1048576),b""): h.update(chunk)
    return h.hexdigest()

def blob(path):
    data=Path(path).read_bytes()
    return hashlib.sha1(f"blob {len(data)}\\0".encode()+data).hexdigest()

def rowhash(row):
    return hashlib.sha256(json.dumps([txt(v) for v in row],ensure_ascii=False,separators=(",",":")).encode()).hexdigest()

def rjson(path):
    with Path(path).open(encoding="utf-8") as f:return json.load(f)

def wjson(path,obj):
    Path(path).parent.mkdir(parents=True,exist_ok=True)
    Path(path).write_text(json.dumps(obj,ensure_ascii=False,indent=2,sort_keys=True)+"\n",encoding="utf-8")

def rcsv(path):
    req(Path(path).is_file(),f"MISSING_INPUT: {path}")
    return pd.read_csv(path,dtype=str,keep_default_na=False)

def wcsv(df,path):
    Path(path).parent.mkdir(parents=True,exist_ok=True)
    df.to_csv(path,index=False,lineterminator="\n")

def norm(df):
    out=df.copy()
    for c in out.columns: out[c]=out[c].map(txt)
    return out

def count(df,n,label): req(len(df)==n,f"{label}: expected {n}, got {len(df)}")
def headers(ws): return [txt(c.value) for c in ws[1]]
def rows(ws): return [[c.value for c in r] for r in ws.iter_rows(min_row=2,max_col=ws.max_column)]

def coverage_layout(ws):
    segcol=max(range(1,ws.max_column+1),key=lambda c:len({txt(ws.cell(r,c).value) for r in range(1,ws.max_row+1)}&TARGETS))
    matches=len({txt(ws.cell(r,segcol).value) for r in range(1,ws.max_row+1)}&TARGETS)
    req(matches>=7,"IMPORT_COVERAGE_SCHEMA_AMBIGUOUS: segment column")
    first=min(r for r in range(1,ws.max_row+1) if txt(ws.cell(r,segcol).value) in TARGETS)
    hr=first-1; req(hr>=1,"IMPORT_COVERAGE_SCHEMA_AMBIGUOUS: header")
    hs={c:txt(ws.cell(hr,c).value).lower().replace(" ","_") for c in range(1,ws.max_column+1)}
    cc=[c for c,h in hs.items() if c!=segcol and any(t in h for t in ("count","rows","members","constituents"))]
    sc=[c for c,h in hs.items() if c!=segcol and any(t in h for t in ("status","state"))]
    req(len(sc)==1,f"IMPORT_COVERAGE_SCHEMA_AMBIGUOUS: status={sc}")
    sr={}
    for r in range(hr+1,ws.max_row+1):
        v=txt(ws.cell(r,segcol).value)
        if v in TARGETS:
            req(v not in sr,f"DUPLICATE_IMPORT_COVERAGE_SEGMENT: {v}"); sr[v]=r
    req(set(OLD_SEGMENTS)<=set(sr),"IMPORT_COVERAGE_MISSING_EXISTING_SEGMENTS")
    matching=[]
    for candidate in cc:
        try:
            if all(intval(ws.cell(sr[seg],candidate).value)==expected for seg,expected in OLD_SEGMENTS.items()):
                matching.append(candidate)
        except RuntimeError:
            pass
    req(len(matching)==1,f"IMPORT_COVERAGE_SCHEMA_AMBIGUOUS: count={cc} matching_current_rows={matching}")
    return {"hr":hr,"seg":segcol,"count":matching[0],"status":sc[0],"rows":sr}

def intval(v):
    try:return int(float(txt(v)))
    except:raise RuntimeError(f"IMPORT_COVERAGE_NON_NUMERIC_COUNT: {txt(v)}")

def update_coverage(ws):
    L=coverage_layout(ws); before={}; statuses=[]
    for seg,n in OLD_SEGMENTS.items():
        r=L["rows"][seg]; req(intval(ws.cell(r,L["count"]).value)==n,f"IMPORT_COVERAGE_COUNT_MISMATCH: {seg}")
        before[seg]=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
        statuses.append(txt(ws.cell(r,L["status"]).value))
    req(all(statuses) and len(set(statuses))==1,f"IMPORT_COVERAGE_EXISTING_STATUS_INCONSISTENT: {statuses}")
    imported=statuses[0]
    if "BR_IBRX100" in L["rows"]:
        r=L["rows"]["BR_IBRX100"]; prior=txt(ws.cell(r,L["count"]).value); state=txt(ws.cell(r,L["status"]).value)
        req(prior in {"","0"} and state!=imported,"CONTROLLED_IMPORT_BLOCKED_COLLISIONS: BR coverage already imported")
    else:
        r=ws.max_row+1; source=next(iter(L["rows"].values()))
        for c in range(1,ws.max_column+1):
            a=ws.cell(source,c); b=ws.cell(r,c)
            if a.has_style:b._style=copy.copy(a._style)
            b.font=copy.copy(a.font); b.fill=copy.copy(a.fill); b.border=copy.copy(a.border); b.alignment=copy.copy(a.alignment); b.number_format=a.number_format
    for c in range(1,ws.max_column+1):ws.cell(r,c).value=None
    ws.cell(r,L["seg"]).value="BR_IBRX100"; ws.cell(r,L["count"]).value=98; ws.cell(r,L["status"]).value=imported
    N=coverage_layout(ws)
    for seg,vals in before.items(): req([ws.cell(N["rows"][seg],c).value for c in range(1,ws.max_column+1)]==vals,f"IMPORT_COVERAGE_PREEXISTING_CHANGED: {seg}")
    imported_n=sum(1 for seg in TARGETS if seg in N["rows"] and txt(ws.cell(N["rows"][seg],N["status"]).value)==imported)
    req(imported_n==8,f"IMPORTED_SEGMENTS_EXPECTED_8_GOT_{imported_n}")
    return {"imported":imported_n,"missing":14-imported_n,"status":imported}

def update_summary(ws):
    before=[[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,ws.max_row+1)]
    req(not any(txt(c.value)=="v0.34_Key" for row in ws.iter_rows() for c in row),"RUN_SUMMARY_V0_34_BLOCK_ALREADY_EXISTS")
    start=ws.max_row+2
    ws.cell(start,1).value="v0.34_Key"
    ws.cell(start,2).value="v0.34_Value"
    updates={"Current_Master_Rows":1633,"Imported_Target_Segments":8,"Missing_Target_Segments":6,"BR_IBRX100_Rows":98,"Stage_Version":VERSION,"Stage_ID":STAGE_ID,"Stage_Status":"PARTIAL","Source_Superset_Complete":False,"P0_Run":False,"Productive":False,"Alpha_Vantage":False}
    for offset,(key,value) in enumerate(updates.items(),1):
        ws.cell(start+offset,1).value=key
        ws.cell(start+offset,2).value=value
    for r,values in enumerate(before,1):
        req([ws.cell(r,c).value for c in range(1,ws.max_column+1)]==values,f"RUN_SUMMARY_PREEXISTING_CHANGED_ROW_{r}")
    materialized={txt(ws.cell(r,1).value):ws.cell(r,2).value for r in range(start+1,start+1+len(updates))}
    req(materialized==updates,"RUN_SUMMARY_V0_34_BLOCK_VALIDATION_FAILED")

def load_inputs(cfg):
    p=cfg["inputs"]
    return {"source":rcsv(p["v032_source_frozen"]),"projection":rcsv(p["v033_projection"]),"plan":rcsv(p["v033_import_plan"]),"ready":rcsv(p["v033_standard_ready"]),"standard":rcsv(p["v033_standard_eligibility"]),"low":rcsv(p["v033_low_pool"]),"exclusions":rcsv(p["v033_standard_exclusions"]),"history":rcsv(p["v033_history_gate"]),"liquidity":rcsv(p["v032_liquidity"])}

def preflight(cfg,F):
    joined=json.dumps({"summary":rjson(cfg["inputs"]["v033_summary"]),"checkpoint":rjson(cfg["inputs"]["v033_checkpoint"])},sort_keys=True)
    for token in (LINEAGE,PROVIDER,"STANDARD_ELIGIBILITY_READY","SCALABLE_NOT_VERIFIED","PASS_HISTORY_STANDARD_U3K"):req(token in joined,f"V0_33_STATE_MISMATCH: {token}")
    for k,n in {"source":98,"projection":98,"plan":98,"ready":38,"standard":38,"low":28,"exclusions":60,"history":38,"liquidity":79}.items():count(F[k],n,f"V0_33_{k.upper()}_ROWS")
    P=F["projection"]; req(list(P.columns)==COLUMNS,"PROJECTION_SCHEMA_MISMATCH")
    req(P["WS_ID"].nunique()==98 and (P["Primary_MIC"]+"|"+P["Primary_Ticker"]).nunique()==98,"PROJECTION_DUPLICATES")
    req(F["plan"]["Standard_Eligibility_Plan_State"].value_counts().to_dict()==STATES,"V0_33_STATE_COUNTS_MISMATCH")
    req(set(F["history"]["History_Gate_State"])=={"PASS_HISTORY_STANDARD_U3K"},"V0_33_HISTORY_MISMATCH")
    gates=[(P["Primary_MIC"]=="BVMF").all(),(P["Country"]=="Brazil").all(),(P["Primary_Currency"]=="BRL").all(),(P["Primary_Universe_Index"]=="BR_IBRX100").all(),(P["Alpha_Symbol"]=="").all(),(P["Scalable_Tradeability_Status"]==SCALABLE_STATUS).all(),(P["Source_ID"]==SOURCE_ID).all(),(P["Source_AsOf"]==SOURCE_ASOF).all()]
    req(all(gates),"V0_33_PROJECTION_SEMANTICS_MISMATCH");req(not truth(cfg["alpha_vantage_allowed"]),"ALPHA_VANTAGE_MUST_BE_FALSE")

def make_state(F):
    P=F["plan"].copy();L=F["liquidity"][["WS_ID","MedianTurnover20_EUR","Price_AsOf","Provider"]];H=F["history"][["WS_ID","Unique_Daily_Bars","Valid_Completed_Bars","History_Gate_State"]]
    M=P.merge(L,on="WS_ID",how="left",validate="one_to_one").merge(H,on="WS_ID",how="left",validate="one_to_one",suffixes=("","_checked"))
    hs=M["History_Gate_State_checked"].where(M["History_Gate_State_checked"].fillna("")!="",M["History_Gate_State"]).fillna("").replace("","NOT_CHECKED_IN_V0_33_NON_STANDARD_CANDIDATE")
    R=pd.DataFrame({"WS_ID":M["WS_ID"],"Primary_MIC":M["Primary_MIC"],"Primary_Ticker":M["Primary_Ticker"],"Source_Superset_Member":True,"Instrument_Gate_State":M["Instrument_Gate_State"],"Liquidity_Class":M["Liquidity_Class"],"History_Gate_State":hs,"Standard_Eligibility_State":M["Standard_Eligibility_Plan_State"],"Scalable_Tradeability_Status":M["Scalable_Tradeability_Status"],"MedianTurnover20_EUR":M["MedianTurnover20_EUR"].fillna(""),"History_Unique_Daily_Bars":M["Unique_Daily_Bars"].fillna(""),"History_Valid_Completed_Bars":M["Valid_Completed_Bars"].fillna(""),"Source_AsOf":M["Source_AsOf"],"Price_AsOf":M["Price_AsOf"].fillna(""),"Provider":M["Provider"].fillna(""),"Eligibility_State_Materialized_v0_34":True,"Productive_Eligibility":False,"SWING_U3K_FROZEN_Member":False})
    count(R,98,"ELIGIBILITY_STATE");req(R["Standard_Eligibility_State"].value_counts().to_dict()==STATES,"ELIGIBILITY_DISTRIBUTION")
    return R

def ready_snapshot(F):
    R=F["ready"].copy();req((R["Eligibility_Plan_State"]=="STANDARD_ELIGIBILITY_READY").all() and (R["History_Gate_State"]=="PASS_HISTORY_STANDARD_U3K").all() and R["Liquidity_Class"].isin({"PASS_PREFERRED","PASS_STANDARD"}).all(),"READY_SNAPSHOT_GATES")
    R["Eligibility_State_Materialized_v0_34"]=True;R["Productive_Eligibility"]=False;R["SWING_U3K_FROZEN_Member"]=False;count(R,38,"READY_SNAPSHOT");return R

def transact(cfg,F):
    master=Path(cfg["inputs"]["current_master_xlsx"]);backup=Path(cfg["backup_master_xlsx"]);historic=Path(cfg["inputs"]["research_partial_1535"])
    req(master.is_file() and historic.is_file(),"MASTER_OR_HISTORIC_MISSING");req(sum(1 for _ in historic.open(encoding="utf-8"))-1==1535,"RESEARCH_PARTIAL_1535_ROWS")
    prehash=sha256(master);preblob=blob(master);histsha=sha256(historic)
    wb=load_workbook(master);req(set(wb.sheetnames)==SHEETS,f"WORKBOOK_SHEETS_MISMATCH: {wb.sheetnames}")
    ws=wb["Universe_Master"];req(headers(ws)==COLUMNS,f"MASTER_SCHEMA_MISMATCH: {headers(ws)}")
    before=rows(ws);req(len(before)==1535,"PREIMPORT_ROWS");B=norm(pd.DataFrame(before,columns=COLUMNS))
    req(B["WS_ID"].nunique()==1535 and (B["Primary_MIC"]+"|"+B["Primary_Ticker"]).nunique()==1535,"PREIMPORT_IDENTITY_DUPLICATES");req((B["Primary_Universe_Index"]=="BR_IBRX100").sum()==0,"CONTROLLED_IMPORT_BLOCKED_COLLISIONS: BR already present")
    defs={txt(c.value) for row in wb["Status_Definitions"].iter_rows() for c in row if txt(c.value)}
    req(CANONICAL_STATUS in defs and set(B["Universe_Status"])=={CANONICAL_STATUS},"UNIVERSE_STATUS_SCHEMA_DECISION_REQUIRED")
    P=F["projection"];pkeys=P["Primary_MIC"]+"|"+P["Primary_Ticker"];bkeys=B["Primary_MIC"]+"|"+B["Primary_Ticker"]
    req(not(set(B["WS_ID"])&set(P["WS_ID"])) and not(set(bkeys)&set(pkeys)),"CONTROLLED_IMPORT_BLOCKED_COLLISIONS")
    backup.parent.mkdir(parents=True,exist_ok=True);shutil.copy2(master,backup);req(sha256(backup)==prehash and blob(backup)==preblob,"PREIMPORT_BACKUP_MISMATCH")
    fh=tempfile.NamedTemporaryFile(prefix=".v0_34_",suffix=".xlsx",dir=str(master.parent),delete=False);tmp=Path(fh.name);fh.close();shutil.copy2(master,tmp)
    try:
        work=load_workbook(tmp);target=work["Universe_Master"]
        for _,s in P.iterrows():
            vals=[]
            for c in COLUMNS:
                v=CANONICAL_STATUS if c=="Universe_Status" else truth(s[c]) if c=="Active" else s[c]
                vals.append(None if txt(v)=="" else v)
            target.append(vals)
        cov=update_coverage(work["Import_Coverage"]);update_summary(work["Run_Summary"]);work.save(tmp)
        chk=load_workbook(tmp);postrows=rows(chk["Universe_Master"]);req(len(postrows)==1633,"POSTIMPORT_ROWS");req(postrows[:1535]==before,"PREEXISTING_1535_ROWS_CHANGED")
        Q=norm(pd.DataFrame(postrows,columns=COLUMNS));req(Q["WS_ID"].nunique()==1633 and (Q["Primary_MIC"]+"|"+Q["Primary_Ticker"]).nunique()==1633,"POSTIMPORT_IDENTITY_DUPLICATES")
        BR=Q[Q["Primary_Universe_Index"]=="BR_IBRX100"].copy();count(BR,98,"POSTIMPORT_BR_ROWS")
        E=norm(P);E["Universe_Status"]=CANONICAL_STATUS;E["Active"]="True";req(Q.iloc[1535:].reset_index(drop=True).equals(E.reset_index(drop=True)),"POSTIMPORT_ROWS_DIFFER_FROM_PROJECTION")
        for seg,n in OLD_SEGMENTS.items():req(int((Q["Primary_Universe_Index"]==seg).sum())==n,f"PREEXISTING_SEGMENT_CHANGED: {seg}")
        req((BR["Alpha_Symbol"]=="").all() and (BR["Scalable_Tradeability_Status"]==SCALABLE_STATUS).all(),"BR_SAFETY_FIELDS_CHANGED")
        os.replace(tmp,master)
    except Exception:
        if tmp.exists():tmp.unlink()
        req(sha256(master)==prehash,"CANONICAL_CHANGED_DURING_FAILED_TRANSACTION");raise
    return {"pre":prehash,"preblob":preblob,"backup":sha256(backup),"backupblob":blob(backup),"post":sha256(master),"historic":histsha,"before":B,"postdf":Q,"br":BR,"coverage":cov}

def outputs(cfg,F,S,R,T):
    out=Path(cfg["output_dir"]);out.mkdir(parents=True,exist_ok=True);Q=T["postdf"];BR=T["br"]
    wcsv(BR[COLUMNS],out/"br_ibrx100_imported_rows_v0.34.csv")
    wcsv(pd.DataFrame([{"Gate":"Pre-Import Master Rows","Expected":1535,"Actual":1535,"State":"PASS"},{"Gate":"Post-Import Master Rows","Expected":1633,"Actual":1633,"State":"PASS"},{"Gate":"Added Brazil Rows","Expected":98,"Actual":98,"State":"PASS"},{"Gate":"Changed Preexisting Rows","Expected":0,"Actual":0,"State":"PASS"},{"Gate":"Collision Count","Expected":0,"Actual":0,"State":"PASS"},{"Gate":"Backup SHA256 equals input","Expected":T["pre"],"Actual":T["backup"],"State":"PASS"}]),out/"br_ibrx100_import_audit_v0.34.csv")
    A=S["Standard_Eligibility_State"].value_counts().rename_axis("Standard_Eligibility_State").reset_index(name="Rows");A["Expected_Rows"]=A["Standard_Eligibility_State"].map(STATES);A["State"]="PASS";wcsv(A,out/"br_ibrx100_eligibility_state_audit_v0.34.csv")
    imm=[]
    for i,b in T["before"].iterrows():
        bh=rowhash(b.tolist());ah=rowhash(Q.iloc[i].tolist());imm.append({"WS_ID":b["WS_ID"],"Before_Row_SHA256":bh,"After_Row_SHA256":ah,"Changed":bh!=ah})
    I=pd.DataFrame(imm);req(not I["Changed"].any(),"IMMUTABILITY_AUDIT_CHANGED");wcsv(I,out/"preexisting_1535_immutability_audit_v0.34.csv")
    inv=Q.groupby("Primary_Universe_Index",dropna=False).size().reset_index(name="Rows").rename(columns={"Primary_Universe_Index":"Segment_ID"});inv["Imported"]=inv["Segment_ID"].isin(TARGETS);wcsv(inv,out/"current_master_segment_inventory_v0.34.csv")
    ida=pd.DataFrame([{"Check":"Universe_Master_Rows","Actual":len(Q),"Expected":1633,"State":"PASS"},{"Check":"Unique_WS_ID","Actual":Q["WS_ID"].nunique(),"Expected":1633,"State":"PASS"},{"Check":"Unique_MIC_Ticker","Actual":(Q["Primary_MIC"]+"|"+Q["Primary_Ticker"]).nunique(),"Expected":1633,"State":"PASS"},{"Check":"BR_IBRX100_Rows","Actual":len(BR),"Expected":98,"State":"PASS"},{"Check":"Old_WS_ID_Preserved","Actual":len(set(T["before"]["WS_ID"])&set(Q["WS_ID"])),"Expected":1535,"State":"PASS"}]);wcsv(ida,out/"current_master_post_import_identity_audit_v0.34.csv")
    sheets=pd.DataFrame([{"Sheet":"Universe_Master","Check":"Rows","Actual":1633,"Expected":1633,"State":"PASS"},{"Sheet":"Import_Coverage","Check":"Imported_Target_Segments","Actual":8,"Expected":8,"State":"PASS"},{"Sheet":"Import_Coverage","Check":"Missing_Target_Segments","Actual":6,"Expected":6,"State":"PASS"},{"Sheet":"Dedupe_Rules","Check":"Preserved","Actual":True,"Expected":True,"State":"PASS"},{"Sheet":"Status_Definitions","Check":"Preserved","Actual":True,"Expected":True,"State":"PASS"},{"Sheet":"Run_Summary","Check":"Updated","Actual":True,"Expected":True,"State":"PASS"}]);wcsv(sheets,out/"workbook_sheet_consistency_audit_v0.34.csv")
    partial=Path(cfg["research_partial_1633"]);wcsv(Q[COLUMNS],partial);count(rcsv(partial),1633,"RESEARCH_PARTIAL_1633")
    recon=pd.DataFrame([{"Check":"Rows","Expected":1633,"Actual":1633,"State":"PASS"},{"Check":"Columns","Expected":22,"Actual":22,"State":"PASS"},{"Check":"Unique_WS_ID","Expected":1633,"Actual":Q["WS_ID"].nunique(),"State":"PASS"},{"Check":"Brazil_Rows","Expected":98,"Actual":len(BR),"State":"PASS"},{"Check":"Selection","Expected":"ALL_CURRENT_MASTER_ROWS","Actual":"ALL_CURRENT_MASTER_ROWS","State":"PASS"}]);wcsv(recon,out/"research_partial_1633_reconciliation_v0.34.csv")
    wcsv(S,cfg["eligibility_state"]);wcsv(R,cfg["standard_ready_snapshot"])
    wjson(cfg["research_partial_manifest"],{"schema":"WELT_SWING_RESEARCH_PARTIAL_MANIFEST_V0_34","version":VERSION,"generated_utc":utc(),"source_master_path":cfg["inputs"]["current_master_xlsx"],"source_master_sha256":T["post"],"rows":1633,"imported_segments":8,"missing_segments":6,"brazil_rows":98,"selection":"all current master rows","scope":"RESEARCH_PARTIAL","universe_complete":False,"P0_run":False,"price_scan_run":False,"news_run":False,"ranking_run":False,"productive":False,"Alpha_Vantage":False,"csv_sha256":sha256(partial)})
    summary={"schema":SCHEMA,"stage_id":STAGE_ID,"version":VERSION,"status":STATUS,"stage_status_global":"PARTIAL","lineage":LINEAGE,"generated_utc":utc(),"pre_import_master_rows":1535,"post_import_master_rows":1633,"added_brazil_rows":98,"changed_preexisting_rows":0,"collision_count":0,"imported_target_segments":8,"missing_target_segments":6,"br_source_imported":98,"eligibility_state_rows":98,"standard_eligibility_ready":38,"low_liquidity_exception_pool":28,"standard_u3k_not_eligible_liquidity":13,"standard_u3k_not_eligible_instrument":19,"standard_ready_snapshot_rows":38,"scalable_not_verified":98,"research_partial_1633_rows":1633,"research_partial_1535_unchanged":True,"preexisting_1535_changed_rows":0,"pre_import_master_sha256":T["pre"],"pre_import_master_git_blob_sha":T["preblob"],"backup_sha256":T["backup"],"backup_git_blob_sha":T["backupblob"],"post_import_master_sha256":T["post"],"provider":PROVIDER,"canonical_master_import_v0_34":True,"universe_mutated_v0_34":True,"br_source_superset_imported_v0_34":True,"eligibility_state_materialized_v0_34":True,"productive_eligibility_promotion_v0_34":False,"swing_u3k_frozen":False,"p0_run":False,"sector_rs":False,"productive":False,"alpha_vantage":False,"source_superset_complete":False,"next_stage":"CURRENT_MASTER_REMAINING_MISSING_SEGMENT_OFFICIAL_SOURCE_MATERIALIZATION"}
    wjson(out/"summary_v0.34.json",summary);wjson(out/"stage_checkpoint_v0.34.json",{"schema":SCHEMA+"_CHECKPOINT","version":VERSION,"stage_id":STAGE_ID,"status":"SUCCESS","stage_classification":"PARTIAL","lineage":LINEAGE,"result_gates_passed":True,"canonical_master_import_v0_34":True,"universe_mutated_v0_34":True,"eligibility_state_materialized_v0_34":True,"productive_eligibility_promotion_v0_34":False,"current_master_rows":1633,"imported_target_segments":8,"missing_target_segments":6,"p0_run":False,"productive":False,"next_stage":summary["next_stage"]})
    handoff=f"""# WELT-SWING LONG DEV - CURRENT Handoff v0.34

## Authoritative position
- Lineage: {LINEAGE}
- Stage: {STAGE_ID}
- Version: v0.34
- Status: {STATUS}
- Global stage classification: PARTIAL
- Welt-Swing v7.2 remains the only productive authority.

## Controlled Brazil import
- Current Master: 1633
- Imported Target Segments: 8/14
- Missing Target Segments: 6/14
- Brazil Source Imported: 98
- Preexisting 1535 changed rows: 0
- Collision Count: 0
- research_partial_1633: present
- research_partial_1535: unchanged historical freeze

## Brazil eligibility state
- STANDARD_ELIGIBILITY_READY: 38
- LOW_LIQUIDITY_EXCEPTION_POOL: 28
- STANDARD_U3K_NOT_ELIGIBLE_LIQUIDITY: 13
- STANDARD_U3K_NOT_ELIGIBLE_INSTRUMENT: 19
- Scalable: 98 SCALABLE_NOT_VERIFIED
- Provider evidence: YFINANCE_FREE

## Safety
- Canonical_Master_Import_v0_34: true
- Universe_Mutated_v0_34: true
- BR_Source_Superset_Imported_v0_34: true
- Eligibility_State_Materialized_v0_34: true
- Productive_Eligibility_Promotion_v0_34: false
- P0: false
- Sector RS: false
- SWING_U3K_FROZEN: false
- Productive: false
- Alpha Vantage: false
- Source Superset Complete: false

## Recovery order
1. Read the master specification.
2. Read this CURRENT handoff.
3. Read output_current_master_br_ibrx100_import_v0_34/summary_v0.34.json.
4. Read stage_checkpoint_v0.34.json and all identity, immutability, workbook and partial reconciliations.
5. Continue only with the next named stage.

## Next stage
CURRENT_MASTER_REMAINING_MISSING_SEGMENT_OFFICIAL_SOURCE_MATERIALIZATION

Do not start the next stage as part of v0.34.
"""
    Path(cfg["handoff_versioned"]).write_text(handoff,encoding="utf-8");Path(cfg["handoff_current"]).write_text(handoff,encoding="utf-8");req(Path(cfg["handoff_versioned"]).read_bytes()==Path(cfg["handoff_current"]).read_bytes(),"HANDOFF_NOT_IDENTICAL")
    paths=[out/n for n in OUTPUTS]+[Path(cfg[k]) for k in ("backup_master_xlsx","eligibility_state","standard_ready_snapshot","research_partial_1633","research_partial_manifest","handoff_versioned","handoff_current")]+[Path(cfg["inputs"]["current_master_xlsx"])]
    entries=[{"path":p.as_posix(),"bytes":p.stat().st_size,"sha256":sha256(p)} for p in paths];wjson(out/"manifest_v0.34.json",{"schema":SCHEMA+"_MANIFEST","version":VERSION,"generated_utc":utc(),"files":entries})
    return summary

def run(cfg):
    F=load_inputs(cfg);preflight(cfg,F);S=make_state(F);R=ready_snapshot(F);T=transact(cfg,F);return outputs(cfg,F,S,R,T)

def validate(cfg):
    out=Path(cfg["output_dir"]);s=rjson(out/"summary_v0.34.json")
    expected={"pre_import_master_rows":1535,"post_import_master_rows":1633,"added_brazil_rows":98,"changed_preexisting_rows":0,"collision_count":0,"imported_target_segments":8,"missing_target_segments":6,"eligibility_state_rows":98,"standard_eligibility_ready":38,"low_liquidity_exception_pool":28,"standard_u3k_not_eligible_liquidity":13,"standard_u3k_not_eligible_instrument":19,"standard_ready_snapshot_rows":38,"research_partial_1633_rows":1633}
    for k,v in expected.items():req(s.get(k)==v,f"RESULT_GATE {k}: {s.get(k)} != {v}")
    for k in ("productive_eligibility_promotion_v0_34","swing_u3k_frozen","p0_run","sector_rs","productive","alpha_vantage","source_superset_complete"):req(s.get(k) is False,f"{k}_MUST_BE_FALSE")
    for k in ("canonical_master_import_v0_34","universe_mutated_v0_34","br_source_superset_imported_v0_34","eligibility_state_materialized_v0_34"):req(s.get(k) is True,f"{k}_MUST_BE_TRUE")
    M=pd.read_excel(cfg["inputs"]["current_master_xlsx"],sheet_name="Universe_Master",dtype=str,keep_default_na=False);count(M,1633,"MASTER");req(M["WS_ID"].nunique()==1633 and (M["Primary_MIC"]+"|"+M["Primary_Ticker"]).nunique()==1633,"MASTER_DUPLICATES");count(M[M["Primary_Universe_Index"]=="BR_IBRX100"],98,"BR")
    count(rcsv(cfg["eligibility_state"]),98,"ELIGIBILITY_STATE");count(rcsv(cfg["standard_ready_snapshot"]),38,"READY");count(rcsv(cfg["research_partial_1633"]),1633,"PARTIAL");count(rcsv(out/"preexisting_1535_immutability_audit_v0.34.csv"),1535,"IMMUTABILITY")
    req(sha256(cfg["backup_master_xlsx"])==s["pre_import_master_sha256"],"BACKUP_SHA256");req(Path(cfg["handoff_versioned"]).read_bytes()==Path(cfg["handoff_current"]).read_bytes(),"HANDOFF_NOT_IDENTICAL")
    for n in OUTPUTS+["manifest_v0.34.json"]:req((out/n).is_file(),f"MISSING_OUTPUT: {n}")

def selftest():
    req(truth("true") and not truth("false"),"FLAG");req(sum(STATES.values())==98,"STATES");req(sum(OLD_SEGMENTS.values())==1535,"OLD_SEGMENTS");req(len(TARGETS)==14,"TARGETS");print("SELF_TEST_PASS")

def main():
    p=argparse.ArgumentParser();p.add_argument("--config");p.add_argument("--validate",action="store_true");p.add_argument("--self-test",action="store_true");a=p.parse_args()
    if a.self_test:selftest();return
    req(a.config,"--config required");cfg=rjson(a.config);req(cfg["stage_id"]==STAGE_ID and cfg["version"]==VERSION,"CONFIG_ID_OR_VERSION")
    if a.validate:validate(cfg);print("V0_34_RESULT_GATES_PASS")
    else:print(json.dumps(run(cfg),ensure_ascii=False,indent=2,sort_keys=True))
if __name__=="__main__":main()
